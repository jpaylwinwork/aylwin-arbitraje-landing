#!/usr/bin/env python3
"""Genera el CSV de importación de conversiones offline de Google Ads desde LEADS_ADS.xlsx.
Uso: python generar_csv_conversiones.py LEADS_ADS.xlsx [salida.csv] [--desde AAAA-MM-DD]
Exporta filas con Click_ID y Estado 3/4/5. Verificar la plantilla vigente en Google Ads antes de importar."""
import sys, csv, datetime, re
from openpyxl import load_workbook

args = [a for a in sys.argv[1:] if not a.startswith("--")]
desde = None
for a in sys.argv[1:]:
    if a.startswith("--desde"):
        desde = datetime.date.fromisoformat(a.split("=",1)[1] if "=" in a else sys.argv[sys.argv.index(a)+1])
xlsx = args[0] if args else "LEADS_ADS.xlsx"
out = args[1] if len(args) > 1 else "conversiones_offline.csv"

wb = load_workbook(xlsx, data_only=True)
cfg = {r[0].value: r[1].value for r in wb["Config"].iter_rows(min_row=2) if r[0].value}
ws = wb["Leads"]
head = [c.value for c in ws[1]]
ix = {h: i for i, h in enumerate(head)}
MAP = {"3": ("Conv_Reunion", "reunion"), "4": ("Conv_Propuesta", "propuesta"), "5": ("Conv_Contratado", "contratado")}

rows, rows_braid, omitidas = [], [], 0
for r in ws.iter_rows(min_row=2, values_only=True):
    estado, click = r[ix["Estado"]], r[ix["Click_ID"]]
    if not estado or not click:
        omitidas += 1 if estado and str(estado)[0] in MAP and not click else 0
        continue
    key = str(estado)[0]
    if key not in MAP: continue
    fecha = r[ix["Fecha_Estado"]] or r[ix["Fecha_Lead"]]
    if isinstance(fecha, str): fecha = datetime.datetime.fromisoformat(fecha)
    if desde and fecha.date() < desde: continue
    conv_name = cfg.get(MAP[key][0])
    if key == "3": valor = cfg.get("Valor_Reunion_CLP") or 0
    else:
        uf = float(r[ix["Honorario_UF"]] or 0) * float(cfg.get("Valor_UF_CLP") or 0)
        valor = round(uf * float(cfg.get("Pct_Valor_Propuesta") or 0.1)) if key == "4" else round(uf)
        if not valor: valor = r[ix["Valor_Conversion_CLP"]] or 0
    cid = str(click).strip()
    tipo = "gclid"
    m = re.match(r"^(gclid|wbraid|gbraid):(.+)$", cid)
    if m: tipo, cid = m.group(1), m.group(2)
    fila = [cid, conv_name, fecha.strftime("%Y-%m-%d %H:%M:%S"), valor, "CLP"]
    (rows if tipo == "gclid" else rows_braid).append(fila)

with open(out, "w", newline="", encoding="utf-8") as f:
    f.write(f"Parameters:TimeZone={cfg.get('TimeZone','America/Santiago')}\n")
    w = csv.writer(f)
    w.writerow(["Google Click ID","Conversion Name","Conversion Time","Conversion Value","Conversion Currency"])
    w.writerows(rows)
if rows_braid:
    out2 = out.replace(".csv", "_braid.csv")
    with open(out2, "w", newline="", encoding="utf-8") as f:
        f.write(f"Parameters:TimeZone={cfg.get('TimeZone','America/Santiago')}\n")
        w = csv.writer(f)
        w.writerow(["wbraid","Conversion Name","Conversion Time","Conversion Value","Conversion Currency"])
        w.writerows(rows_braid)
    print(f"{out2}: {len(rows_braid)} conversiones iOS (wbraid/gbraid). Importarlas por separado: la plantilla de Google para wbraid/gbraid usa otra columna. Verificar plantilla vigente en la interfaz.")
print(f"{out}: {len(rows)} conversiones GCLID exportadas." + (f" AVISO: {omitidas} filas en estado 3/4/5 sin Click_ID (no importables)." if omitidas else ""))
