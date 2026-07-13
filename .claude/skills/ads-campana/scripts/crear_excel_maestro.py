#!/usr/bin/env python3
"""Crea LEADS_ADS.xlsx, el Excel maestro de leads de Google Ads de Aylwin Matta.
Uso: python crear_excel_maestro.py [ruta_salida.xlsx]
No sobreescribe si el archivo ya existe."""
import sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

PATH = sys.argv[1] if len(sys.argv) > 1 else "LEADS_ADS.xlsx"
if os.path.exists(PATH):
    sys.exit(f"ERROR: {PATH} ya existe. No se sobreescribe el Excel maestro.")

AZUL, BLANCO = "1F3864", "FFFFFF"
COLS = ["N","Fecha_Lead","Canal","Campana","Grupo","Keyword","Click_ID","Nombre","Empresa",
        "Telefono","Email","Materia","Resumen","Estado","Fecha_Estado","Honorario_UF",
        "Valor_Conversion_CLP","Notas"]
ESTADOS = '"1.Lead nuevo,2.Calificado,3.Reunión realizada,4.Propuesta enviada,5.Contratado,0.Descartado"'
CANALES = '"WhatsApp,Formulario,Llamada,Otro"'
MATERIAS = '"Arbitraje,Reclamo de ilegalidad,Otra"'

wb = Workbook()
ws = wb.active; ws.title = "Leads"
for i, c in enumerate(COLS, 1):
    cell = ws.cell(1, i, c)
    cell.font = Font(bold=True, color=BLANCO); cell.fill = PatternFill("solid", fgColor=AZUL)
    cell.alignment = Alignment(horizontal="center")
ws.freeze_panes = "A2"
for col, width in zip("ABCDEFGHIJKLMNOPQR", [4,16,11,18,16,22,28,18,20,14,24,20,40,18,12,12,18,30]):
    ws.column_dimensions[col].width = width
for formula, rng in [(ESTADOS,"N2:N500"),(CANALES,"C2:C500"),(MATERIAS,"L2:L500")]:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv); dv.add(rng)

cfg = wb.create_sheet("Config")
cfg.append(["Parametro","Valor","Nota"])
for r in [
    ("TimeZone","America/Santiago","Para el CSV de importación"),
    ("Moneda","CLP",""),
    ("Conv_Lead_WhatsApp","Lead_WhatsApp","Nombre EXACTO en Google Ads"),
    ("Conv_Lead_Formulario","Lead_Formulario","Nombre EXACTO en Google Ads"),
    ("Conv_Reunion","Reunion_Calificada","Nombre EXACTO en Google Ads"),
    ("Conv_Propuesta","Propuesta_Enviada","Nombre EXACTO en Google Ads"),
    ("Conv_Contratado","Caso_Contratado","Nombre EXACTO en Google Ads"),
    ("Valor_Reunion_CLP",50000,"Valor fijo por reunión calificada"),
    ("Pct_Valor_Propuesta",0.10,"Fracción del honorario propuesto"),
    ("Valor_UF_CLP",0,"Actualizar manualmente al valor vigente"),
]: cfg.append(list(r))
for i in range(1,4): cfg.cell(1,i).font = Font(bold=True)
cfg.column_dimensions["A"].width = 26; cfg.column_dimensions["B"].width = 22; cfg.column_dimensions["C"].width = 40

wb.save(PATH)
print(f"Creado: {PATH}")
